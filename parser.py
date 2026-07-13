import os
import openpyxl
import pandas as pd
from datetime import datetime

def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def parse_single_crf(file_path_or_bytes, filename="Unknown"):
    """
    1つのExcelファイルをパースし、各シートのデータを構造化した辞書として返す。
    数式のキャッシュ値が読み込めない場合は、登録票から取得した基本情報で補完する。
    """
    try:
        # data_only=True で数式の計算結果を取得
        wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True)
    except Exception as e:
        print(f"Error loading workbook {filename}: {e}")
        return None

    # --- 1. 登録票のパース (患者基本情報の取得) ---
    reg_info = {}
    if "登録票" in wb.sheetnames:
        ws = wb["登録票"]
        current_category = ""
        for r in range(1, ws.max_row + 1):
            cat_val = ws.cell(r, 1).value
            item_val = ws.cell(r, 2).value
            val = ws.cell(r, 3).value
            
            if cat_val is not None:
                current_category = str(cat_val).strip()
            
            if item_val is not None:
                item_str = str(item_val).strip()
                key = f"{current_category}_{item_str}" if current_category else item_str
                reg_info[key] = val

    # 基本情報の特定（他のシートの補完に使用）
    # キー名は '患者情報_カルテ番号' や '患者情報_性別' など
    karte_no = None
    gender = None
    birth_date = None
    
    for k, v in reg_info.items():
        if "カルテ番号" in k:
            karte_no = safe_str(v)
        elif "性別" in k:
            gender = safe_str(v)
        elif "生年月日" in k:
            birth_date = v

    if not karte_no:
        # フォールバック: ファイル名からカルテ番号を推測するか、デフォルト値を設定
        karte_no = f"K-{filename.replace('.xlsx', '')}"
    
    # 全体の辞書に基本情報を設定
    meta = {
        "ファイル名": filename,
        "カルテ番号": karte_no,
        "性別": gender,
        "生年月日": birth_date
    }

    # 登録票のデータをフラットな辞書にする
    flat_reg = {**meta}
    for k, v in reg_info.items():
        flat_reg[k] = v

    # --- 2. 登録時情報のパース (診断・治療歴 ＋ その他基本) ---
    diagnoses = []
    baseline_info = {**meta}
    
    if "登録時情報" in wb.sheetnames:
        ws = wb["登録時情報"]
        
        # 診断・治療歴テーブルの取得
        # 3行目から開始し、次のセクションが始まるか空行になるまで
        r = 3
        while r <= ws.max_row:
            disease = ws.cell(r, 1).value
            if disease == "化学予防" or disease is None:
                break
            
            diag = {
                "ファイル名": filename,
                "カルテ番号": karte_no,
                "病名": safe_str(disease),
                "原発部位": safe_str(ws.cell(r, 2).value),
                "サブサイト": safe_str(ws.cell(r, 3).value),
                "組織型": safe_str(ws.cell(r, 4).value),
                "診断年齢": ws.cell(r, 5).value,
                "治療内容": safe_str(ws.cell(r, 6).value)
            }
            diagnoses.append(diag)
            r += 1

        # その他基本セクションのパース
        r = 1
        while r <= ws.max_row:
            val_a = ws.cell(r, 1).value
            val_b = ws.cell(r, 2).value
            val_c = ws.cell(r, 3).value
            
            # 妊娠・出産情報
            if val_a == "妊娠・出産情報" and val_b == "妊娠回数":
                baseline_info["妊娠回数"] = val_c
            elif val_a is None and val_b == "出産回数":
                baseline_info["出産回数"] = val_c
                
            # 身長・体重
            elif val_a == "身長・体重":
                h_row = r + 2
                if h_row <= ws.max_row:
                    baseline_info["身長_cm"] = ws.cell(h_row, 1).value
                    baseline_info["体重_kg"] = ws.cell(h_row, 2).value
                    baseline_info["BMI"] = ws.cell(h_row, 3).value
                    
            # 喫煙歴
            elif val_a == "喫煙歴":
                s_row = r + 2
                if s_row <= ws.max_row:
                    baseline_info["喫煙_本数_日"] = ws.cell(s_row, 1).value
                    baseline_info["喫煙_開始年齢"] = ws.cell(s_row, 2).value
                    baseline_info["喫煙_禁煙年齢"] = ws.cell(s_row, 3).value
                    
            # 飲酒歴
            elif val_a == "飲酒歴":
                d_row = r + 1
                if d_row <= ws.max_row:
                    baseline_info["飲酒歴"] = safe_str(ws.cell(d_row, 1).value)
                    
            r += 1

    # --- 3. 家系情報のパース ---
    family_history = []
    if "家系情報" in wb.sheetnames:
        ws = wb["家系情報"]
        # 2列目以降が各家族メンバー
        for col in range(2, ws.max_column + 1):
            relation = ws.cell(2, col).value
            if relation is None or safe_str(relation) == "例":
                continue
            
            f_gender = safe_str(ws.cell(3, col).value)
            outcome = safe_str(ws.cell(4, col).value)
            
            # 疾患１〜５
            diseases = []
            for i in range(5, 15, 2):
                dis_name = ws.cell(i, col).value
                dis_age = ws.cell(i+1, col).value
                if dis_name:
                    diseases.append({"病名": safe_str(dis_name), "診断年齢": dis_age})
            
            family_history.append({
                "ファイル名": filename,
                "カルテ番号": karte_no,
                "関係": safe_str(relation),
                "性別": f_gender,
                "転帰": outcome,
                "疾患一覧": diseases # エラーチェック用のリスト
            })

    # --- 4. サーベイランスのパース ---
    surveillances = []
    if "サーベイランス" in wb.sheetnames:
        ws = wb["サーベイランス"]
        # 5行目以降がデータ
        for r in range(5, ws.max_row + 1):
            exam_type = ws.cell(r, 1).value
            exam_date = ws.cell(r, 2).value
            if exam_type is None and exam_date is None:
                continue
            
            surveillances.append({
                "ファイル名": filename,
                "カルテ番号": karte_no,
                "検査種別": safe_str(exam_type),
                "検査日": exam_date,
                "検査所見": safe_str(ws.cell(r, 3).value),
                "確認検査": safe_str(ws.cell(r, 4).value),
                "確認検査日": ws.cell(r, 5).value,
                "病理診断方法": safe_str(ws.cell(r, 6).value),
                "ICD10": safe_str(ws.cell(r, 7).value),
                "OncoTree_LV1": safe_str(ws.cell(r, 8).value),
                "OncoTree": safe_str(ws.cell(r, 9).value),
                "治療内容": safe_str(ws.cell(r, 10).value),
                "治療内容_自由記載": safe_str(ws.cell(r, 11).value)
            })

    # --- 5. 化学予防のパース ---
    chemoprevention = []
    if "化学予防" in wb.sheetnames:
        ws = wb["化学予防"]
        for r in range(5, ws.max_row + 1):
            drug_name = ws.cell(r, 2).value
            start_date = ws.cell(r, 3).value
            if drug_name is None and start_date is None:
                continue
            
            chemoprevention.append({
                "ファイル名": filename,
                "カルテ番号": karte_no,
                "薬剤名": safe_str(drug_name),
                "内服開始日": start_date,
                "内服終了日": ws.cell(r, 4).value,
                "終了理由": safe_str(ws.cell(r, 5).value),
                "有害事象_1": safe_str(ws.cell(r, 6).value),
                "Grade_1": safe_str(ws.cell(r, 7).value),
                "有害事象_2": safe_str(ws.cell(r, 8).value),
                "Grade_2": safe_str(ws.cell(r, 9).value),
            })

    # --- 6. 嗜好・生活のパース ---
    lifestyle = {**meta}
    if "嗜好・生活" in wb.sheetnames:
        ws = wb["嗜好・生活"]
        
        # 喫煙歴
        lifestyle["喫煙歴"] = safe_str(ws.cell(5, 2).value)
        lifestyle["喫煙_本数_日"] = ws.cell(5, 3).value
        lifestyle["喫煙_開始年齢"] = ws.cell(5, 4).value
        lifestyle["喫煙_禁煙年齢"] = ws.cell(5, 5).value
        
        # 飲酒歴
        lifestyle["飲酒歴"] = safe_str(ws.cell(7, 2).value)
        
        # 婚姻歴 (結婚_1〜3)
        for i, row_idx in enumerate([9, 11, 13], 1):
            lifestyle[f"結婚_{i}"] = safe_str(ws.cell(row_idx, 2).value)
            lifestyle[f"結婚_{i}_年齢"] = ws.cell(row_idx, 3).value
            lifestyle[f"離婚_{i}"] = safe_str(ws.cell(row_idx, 4).value)
            lifestyle[f"離婚_{i}_年齢"] = ws.cell(row_idx, 5).value

    # --- 7. 転帰・予後のパース ---
    prognosis = {**meta}
    if "転帰・予後" in wb.sheetnames:
        ws = wb["転帰・予後"]
        prognosis["転帰"] = safe_str(ws.cell(4, 3).value)
        prognosis["転院先医療機関"] = safe_str(ws.cell(5, 3).value)
        prognosis["最終生存確認日"] = ws.cell(6, 3).value
        prognosis["死亡年月日"] = ws.cell(7, 3).value

    return {
        "登録票": flat_reg,
        "登録時情報_基本": baseline_info,
        "登録時情報_診断治療歴": diagnoses,
        "家系情報": family_history,
        "サーベイランス": surveillances,
        "化学予防": chemoprevention,
        "嗜好・生活": lifestyle,
        "転帰・予後": prognosis
    }

def combine_multiple_files(uploaded_files_dict):
    """
    アップロードされた複数のファイルを読み込み、シートごとに結合された DataFrame の辞書を返す。
    uploaded_files_dict: { filename: file_path_or_bytes } の辞書
    """
    combined_data = {
        "登録票": [],
        "登録時情報_基本": [],
        "登録時情報_診断治療歴": [],
        "家系情報": [],
        "サーベイランス": [],
        "化学予防": [],
        "嗜好・生活": [],
        "転帰・予後": []
    }
    
    raw_results = []
    
    for filename, file_data in uploaded_files_dict.items():
        parsed = parse_single_crf(file_data, filename)
        if parsed is None:
            continue
        
        raw_results.append(parsed)
        
        for key in combined_data.keys():
            if isinstance(parsed[key], list):
                combined_data[key].extend(parsed[key])
            elif isinstance(parsed[key], dict):
                combined_data[key].append(parsed[key])

    # DataFrame に変換
    dfs = {}
    for key, val_list in combined_data.items():
        if val_list:
            if key == "家系情報":
                flat_fam = []
                for fam in val_list:
                    fam_copy = {k: v for k, v in fam.items() if k != "疾患一覧"}
                    dis_strs = [f"{d['病名']}({d['診断年齢']}歳)" for d in fam["疾患一覧"]]
                    fam_copy["登録疾患"] = ", ".join(dis_strs) if dis_strs else "なし"
                    flat_fam.append(fam_copy)
                dfs[key] = pd.DataFrame(flat_fam)
            else:
                dfs[key] = pd.DataFrame(val_list)
        else:
            dfs[key] = pd.DataFrame()
            
    return dfs, raw_results
